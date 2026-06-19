"""
DocuSense AI — excel_builder.py

Builds the 5-sheet Excel report (Summary Dashboard, All Documents, Line
Items, Cross-Doc Matches, Issues & Flags) from extracted documents and
cross-document analysis results, using openpyxl with the house style:
dark-blue bold headers, alternating row shading, severity-colored issue
rows, Arial 10pt throughout.

Every sheet-building function tolerates empty input (no documents, no
matches, no issues) so a partial run — where some files failed — still
produces a complete, downloadable workbook.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from analyzer import get_doc_amount, get_doc_number, get_vendor, normalize_amount

FONT_NAME = "Arial"
FONT_SIZE = 10

HEADER_FILL = PatternFill(start_color="FF1F4E79", end_color="FF1F4E79", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FFFFFFFF")
ALT_FILL = PatternFill(start_color="FFDEEAF1", end_color="FFDEEAF1", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
BODY_FONT = Font(name=FONT_NAME, size=FONT_SIZE)
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color="FF1F4E79")
SECTION_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FF1F4E79")

SEVERITY_FILLS = {
    "High": PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
    "Medium": PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid"),
    "Low": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
}
SEVERITY_FONTS = {
    "High": Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FFFFFFFF"),
    "Medium": Font(name=FONT_NAME, size=FONT_SIZE, bold=True),
    "Low": Font(name=FONT_NAME, size=FONT_SIZE),
}

WRAP_TOP_LEFT = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# Generic styling helpers
# ---------------------------------------------------------------------------

def _write_header_row(ws: Worksheet, row: int, headers: list[str], start_col: int = 1) -> None:
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def _to_cell_value(value: Any) -> Any:
    """openpyxl can only write str/int/float/bool/datetime/None — flatten anything else (e.g. list-type custom fields)."""
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(v) for v in value)
    if isinstance(value, dict):
        return str(value)
    return value


def _write_data_row(ws: Worksheet, row: int, values: list[Any], alt: bool, start_col: int = 1) -> None:
    fill = ALT_FILL if alt else WHITE_FILL
    for i, value in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=_to_cell_value(value))
        cell.fill = fill
        cell.font = BODY_FONT
        cell.alignment = WRAP_TOP_LEFT


def _write_table(ws: Worksheet, start_row: int, headers: list[str], rows: list[list[Any]], start_col: int = 1) -> int:
    """Writes a header row + alternating data rows starting at start_row. Returns the next free row."""
    _write_header_row(ws, start_row, headers, start_col)
    row = start_row + 1
    for i, values in enumerate(rows):
        _write_data_row(ws, row, values, alt=(i % 2 == 0), start_col=start_col)
        row += 1
    if not rows:
        row += 1  # leave a visible empty row under the header so empty tables aren't confusing
    return row


def _section_title(ws: Worksheet, row: int, text: str) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    return row + 1


def _autosize_columns(ws: Worksheet, max_col: int, min_width: int = 12, max_width: int = 55) -> None:
    for col_idx in range(1, max_col + 1):
        longest = min_width
        for cell in ws[get_column_letter(col_idx)]:
            if cell.value is None:
                continue
            longest = max(longest, min(len(str(cell.value)), max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(longest + 2, max_width)


def _ok_docs(documents: list[dict]) -> list[dict]:
    return [d for d in documents if d.get("success") and isinstance(d.get("data"), dict)]


# ---------------------------------------------------------------------------
# Sheet 1 — Summary Dashboard
# ---------------------------------------------------------------------------

def _build_summary_dashboard(wb: Workbook, documents: list[dict], analysis: dict, generated_at: datetime) -> None:
    ws = wb.create_sheet("Summary Dashboard")
    ok_docs = _ok_docs(documents)
    failed_docs = [d for d in documents if not d.get("success")]
    issues = analysis.get("issues", [])
    vendor_spend = analysis.get("vendor_spend", [])

    ws.cell(row=1, column=1, value="DocuSense AI — Summary Dashboard").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Processed: {generated_at.strftime('%Y-%m-%d %H:%M:%S')}").font = BODY_FONT

    row = 4
    row = _section_title(ws, row, "Documents by Type")
    by_type: dict[str, dict[str, dict[str, float]]] = {}
    for d in ok_docs:
        amount = get_doc_amount(d["doc_type"], d["data"]) or 0.0
        currency = (d["data"].get("currency") or "—").strip() or "—"
        bucket = by_type.setdefault(d["doc_type"], {})
        entry = bucket.setdefault(currency, {"count": 0, "total": 0.0})
        entry["count"] += 1
        entry["total"] += amount

    type_rows = []
    for doc_type, currencies in sorted(by_type.items()):
        for currency, entry in sorted(currencies.items()):
            type_rows.append([doc_type, entry["count"], round(entry["total"], 2), currency])
    row = _write_table(ws, row, ["Document Type", "Count", "Total Value", "Currency"], type_rows)
    row += 1

    row = _section_title(ws, row, "Spend by Vendor")
    vendor_rows = [
        [v["vendor"], v["currency"], round(v["invoice_total"], 2), round(v["po_total"], 2), v["document_count"]]
        for v in vendor_spend
    ]
    row = _write_table(ws, row, ["Vendor", "Currency", "Invoiced Total", "PO Total", "Documents"], vendor_rows)
    row += 1

    row = _section_title(ws, row, "Issues Summary")
    severity_counts = {"High": 0, "Medium": 0, "Low": 0}
    for issue in issues:
        severity_counts[issue.get("severity", "Low")] = severity_counts.get(issue.get("severity", "Low"), 0) + 1
    issue_rows = [[sev, count] for sev, count in severity_counts.items()]
    row = _write_table(ws, row, ["Severity", "Count"], issue_rows)
    row += 1

    row = _section_title(ws, row, "Overall Stats")
    ai_status = "Success" if analysis.get("ai_enrichment_success") else f"Failed ({analysis.get('ai_error')})"
    stat_rows = [
        ["Total Files Uploaded", len(documents)],
        ["Successful Extractions", len(ok_docs)],
        ["Failed Extractions", len(failed_docs)],
        ["Total Issues Found", len(issues)],
        ["AI Cross-Document Enrichment", ai_status],
        ["Processing Date", generated_at.strftime("%Y-%m-%d")],
        ["Processing Time", generated_at.strftime("%H:%M:%S")],
    ]
    row = _write_table(ws, row, ["Metric", "Value"], stat_rows)

    _autosize_columns(ws, 5)


# ---------------------------------------------------------------------------
# Sheet 2 — All Documents
# ---------------------------------------------------------------------------

def _build_key_info(doc_type: str, data: dict) -> str:
    parts = []
    if data.get("status"):
        parts.append(f"Status: {data['status']}")
    if doc_type == "Invoice" and data.get("payment_terms"):
        parts.append(f"Terms: {data['payment_terms']}")
    if doc_type == "Invoice" and data.get("due_date"):
        parts.append(f"Due: {data['due_date']}")
    if doc_type == "Purchase Order" and data.get("approval_status"):
        parts.append(f"Approval: {data['approval_status']}")
    if doc_type == "Contract" and data.get("jurisdiction"):
        parts.append(f"Jurisdiction: {data['jurisdiction']}")
    if doc_type == "Contract" and data.get("expiry_date"):
        parts.append(f"Expires: {data['expiry_date']}")
    key_dates = data.get("key_dates")
    if isinstance(key_dates, list) and key_dates:
        labels = ", ".join(
            f"{kd.get('label', '')}: {kd.get('date', '')}" for kd in key_dates if isinstance(kd, dict)
        )
        if labels:
            parts.append(f"Key dates: {labels}")
    return " | ".join(parts)


def _build_all_documents_sheet(wb: Workbook, documents: list[dict], custom_field_names: list[str] | None = None) -> None:
    custom_field_names = custom_field_names or []
    ws = wb.create_sheet("All Documents")
    headers = ["File Name", "Doc Type", "Doc Number", "Date", "Party 1", "Party 2", "Amount", "Currency", "Key Info", "AI Summary"] + custom_field_names

    rows = []
    for d in documents:
        if d.get("success") and isinstance(d.get("data"), dict):
            data = d["data"]
            row = [
                d["filename"],
                d["doc_type"],
                get_doc_number(d["doc_type"], data) or "",
                data.get("document_date") or "",
                data.get("party_name_1") or "",
                data.get("party_name_2") or "",
                get_doc_amount(d["doc_type"], data),
                data.get("currency") or "",
                _build_key_info(d["doc_type"], data),
                data.get("summary") or "",
            ]
            row.extend(data.get(name) for name in custom_field_names)
            rows.append(row)
        else:
            row = [d["filename"], d.get("doc_type", ""), "", "", "", "", None, "", "", f"ERROR: {d.get('error', 'Extraction failed')}"]
            row.extend([""] * len(custom_field_names))
            rows.append(row)

    _write_table(ws, 1, headers, rows)
    ws.freeze_panes = "A2"
    _autosize_columns(ws, len(headers))


# ---------------------------------------------------------------------------
# Sheet 3 — Line Items
# ---------------------------------------------------------------------------

def _build_line_items_sheet(wb: Workbook, documents: list[dict]) -> None:
    ws = wb.create_sheet("Line Items")
    headers = ["Source Doc", "Doc Number", "Line#", "Description", "Qty", "Unit Price", "Tax", "Total"]

    rows = []
    for d in _ok_docs(documents):
        if d["doc_type"] not in ("Invoice", "Purchase Order"):
            continue
        data = d["data"]
        line_items = data.get("line_items")
        if not isinstance(line_items, list):
            continue
        doc_number = get_doc_number(d["doc_type"], data) or ""
        for idx, item in enumerate(line_items, start=1):
            if not isinstance(item, dict):
                continue
            rows.append([
                d["filename"],
                doc_number,
                idx,
                item.get("description") or "",
                normalize_amount(item.get("quantity")),
                normalize_amount(item.get("unit_price")),
                normalize_amount(item.get("tax")),
                normalize_amount(item.get("total")),
            ])

    _write_table(ws, 1, headers, rows)
    ws.freeze_panes = "A2"
    _autosize_columns(ws, len(headers))


# ---------------------------------------------------------------------------
# Sheet 4 — Cross-Doc Matches
# ---------------------------------------------------------------------------

def _build_cross_doc_matches_sheet(wb: Workbook, matches: list[dict]) -> None:
    ws = wb.create_sheet("Cross-Doc Matches")
    headers = ["PO Number", "PO Amount", "Invoice Number", "Invoice Amount", "Match Status", "Difference", "Flag"]

    rows = [
        [m.get("po_number", ""), m.get("po_amount"), m.get("invoice_number", ""), m.get("invoice_amount"),
         m.get("match_status", ""), m.get("difference"), m.get("flag", "")]
        for m in matches
    ]

    _write_table(ws, 1, headers, rows)
    ws.freeze_panes = "A2"
    _autosize_columns(ws, len(headers))


# ---------------------------------------------------------------------------
# Sheet 5 — Issues & Flags
# ---------------------------------------------------------------------------

def _build_issues_sheet(wb: Workbook, issues: list[dict]) -> None:
    ws = wb.create_sheet("Issues & Flags")
    headers = ["Severity", "Issue Type", "Document(s) Affected", "Details", "Recommended Action"]
    _write_header_row(ws, 1, headers)

    row = 2
    for issue in issues:
        severity = issue.get("severity", "Low")
        fill = SEVERITY_FILLS.get(severity, SEVERITY_FILLS["Low"])
        font = SEVERITY_FONTS.get(severity, SEVERITY_FONTS["Low"])
        values = [severity, issue.get("issue_type", ""), issue.get("documents_affected", ""), issue.get("details", ""), issue.get("recommended_action", "")]
        for i, value in enumerate(values):
            cell = ws.cell(row=row, column=1 + i, value=value)
            cell.fill = fill
            cell.font = font
            cell.alignment = WRAP_TOP_LEFT
        row += 1

    if not issues:
        ws.cell(row=2, column=1, value="No issues detected.").font = BODY_FONT

    ws.freeze_panes = "A2"
    _autosize_columns(ws, len(headers))


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def build_excel_report(documents: list[dict], analysis: dict, custom_field_names: list[str] | None = None) -> BytesIO:
    """
    Build the full 5-sheet DocuSense AI Excel report.

    documents: list of {filename, doc_type, success, data, error}
    analysis: output of analyzer.run_cross_document_analysis
    custom_field_names: optional user-defined field names to add as extra columns in "All Documents"
    """
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet; we add our own in order

    generated_at = datetime.now()

    _build_summary_dashboard(wb, documents, analysis, generated_at)
    _build_all_documents_sheet(wb, documents, custom_field_names=custom_field_names)
    _build_line_items_sheet(wb, documents)
    _build_cross_doc_matches_sheet(wb, analysis.get("matches", []))
    _build_issues_sheet(wb, analysis.get("issues", []))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
